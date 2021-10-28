from openpyxl import load_workbook, workbook
from openpyxl import Workbook
import re

def Results_Sum(wb, ws, name):
    j = 0
    for sheet in wb:
        print()
        i = 0
        if sheet.title == 'Case need update':
            break
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[0] is None:
                ws.append(row)
                i = i + 1
                j = j + 1
        print(sheet.title + ': ' + str(i))
    print('======================== ' + name + ' Total:' + str(j) + ' ========================')


#--------------------------------------------- Input Settings -------------------------------------------------
Week = 43
folder = 'C:\\Users\\Yvonne\\Documents\\Results\\'
Order = ['Main', 'STR','SORP']
Auto = [1, 1, 0]
#--------------------------------------------------------------------------------------------------------------


output = Workbook()
for k in range(len(Order)):
    print(k)
    Line_Name = Order[k]
    Title = 'W' + str(Week) + '_' + Line_Name
    wb =  load_workbook(folder + '\\Sorted\\' + Title +'_Sorted.xlsx')
    ws = output.create_sheet(Line_Name)
    row_title = ['Original GM TC ID', Title]
    ws.append(row_title)
    Results_Sum(wb, ws, Title + '_Man')

    if Auto[k] == 1:
        wa =  load_workbook(folder + '\\Sorted\\' + Title +'_Auto.xlsx')
        Results_Sum(wa, ws, Title + '_Auto')

output.save(folder + '\\All\\2021_W' + str(Week) + '.xlsx')
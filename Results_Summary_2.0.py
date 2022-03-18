from openpyxl import load_workbook, workbook
from openpyxl import Workbook
import re
import glob

#--------------------------------------------- Input Settings -------------------------------------------------
year = 2022
Week = 11
folder = 'C:\\Users\\Yvonne\\Documents\\Results'
Order = ['SORP_Resume','Main_Cold']
#--------------------------------------------------------------------------------------------------------------

def Results_Sum(wb, ws, path):
    j = 0
    for sheet in wb:
        i = 0
        if sheet.title == 'Case need update'or sheet.title == 'Summary' or sheet.title == 'summary' or sheet.title == 'Sheet':
            break
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[0] is None:
                ws.append(row)
                i = i + 1
                j = j + 1
        print(sheet.title + ': ' + str(i))
    print('====================================================================================='\
     + '\nFrom ' + path\
     + '\nTotal:' + str(j)\
     + '\n=====================================================================================\n')

output = Workbook() # Result Summary
for Line_Name in Order:
    print('# ' + Line_Name + ' #')
    Title = 'W' + str(Week).zfill(2) + '_' + Line_Name
    ws = output.create_sheet(Line_Name) # create sheet for each line
    row_title = ['Original GM TC ID', Title] 
    ws.append(row_title) # add title

    files = glob.glob(folder + '\\Sorted\\' + Title +'**.xlsx', recursive = True)
    for xls in files :
        wb =  load_workbook(xls)
        Results_Sum(wb, ws, xls)
output.save(folder + '\\All\\' + str(year) + '_W' + str(Week).zfill(2) + '.xlsx')